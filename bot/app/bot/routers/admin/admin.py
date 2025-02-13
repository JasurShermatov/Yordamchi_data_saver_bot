import os

import pandas as pd
from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message, FSInputFile
from datetime import datetime, timedelta
from aiogram.enums import ChatAction
from openpyxl.styles import PatternFill, Font

from app.bot.constants.admin import (
    STATISTICS,
    USERS_EXCEL,
    CHANNEL_CONTROLLER,
    BACK,
    DATA_CONTROLLER,
    CATEGORY_CONTROLLER,
)
from app.bot.filters.admin import AdminFilter
from app.bot.handlers.admin.statistics import (
    count_users,
    count_users_by_date,
    count_users_between,
)
from app.bot.handlers.admin.users import get_all_users
from app.bot.keyboards.reply.admin import menu_button
from app.bot.keyboards.reply.category import category_button
from app.bot.keyboards.reply.channels import channels_button
from app.bot.keyboards.reply.data import data_button

router = Router()


@router.message(AdminFilter(), Command("admin"))
async def start_handler(message: Message):
    await message.answer(
        "✅ *Admin panelga xush kelibsiz!*\n\n"
        "Bu yerda siz foydalanuvchilarni boshqarish, kategoriyalar qo‘shish va "
        "test ma’lumotlarini tahrirlash imkoniyatiga egasiz.\n\n"
        "⚡ *Tezkor buyruqlar:* \n"
        "➖ Foydalanuvchilar ro‘yxati\n"
        "➖ Kategoriyalarni boshqarish\n"
        "➖ Sozlamalar\n\n"
        "🛠 Kerakli bo‘limni tanlash uchun tugmalardan foydalaning!",
        reply_markup=await menu_button(),
        parse_mode="Markdown",
    )


@router.message(AdminFilter(), F.text == STATISTICS)
async def show_statistics(message: Message):
    await message.bot.send_chat_action(message.chat.id, ChatAction.TYPING)
    try:
        today = datetime.now().date()
        yesterday = today - timedelta(days=1)
        three_days_ago = today - timedelta(days=3)
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        await message.bot.send_chat_action(message.chat.id, ChatAction.TYPING)
        total_users = await count_users()
        today_users = await count_users_by_date(today)
        yesterday_users = await count_users_by_date(yesterday)
        three_days_users = await count_users_between(three_days_ago, today)
        weekly_users = await count_users_between(week_ago, today)
        monthly_users = await count_users_between(month_ago, today)

        stats = [
            "📊 *Bot statistikasi*",
            f"\n📅 *Bugun ({today.strftime('%d.%m.%Y')})* — {today_users:,} ta",
            f"📅 *Kecha ({yesterday.strftime('%d.%m.%Y')})* — {yesterday_users:,} ta",
            f"📅 *Oxirgi 3 kun ({three_days_ago.strftime('%d.%m.%Y')})* — {three_days_users:,} ta",
            f"📅 *Oxirgi 7 kun ({week_ago.strftime('%d.%m.%Y')})* — {weekly_users:,} ta",
            f"📅 *Oxirgi 30 kun ({month_ago.strftime('%d.%m.%Y')})* — {monthly_users:,} ta",
            f"\n🔹 *Jami foydalanuvchilar:* {total_users:,} ta.",
        ]

        await message.answer("\n".join(stats), parse_mode="Markdown")

    except Exception as e:
        print(f"Error showing statistics: {e}")
        await message.answer(
            "❌ *Statistikani olishda xatolik yuz berdi!*", parse_mode="Markdown"
        )


@router.message(AdminFilter(), F.text == USERS_EXCEL)
async def get_users_excel(message: Message):
    await message.answer("📊 Excel fayl tayyorlanmoqda...")

    try:
        await message.bot.send_chat_action(message.chat.id, ChatAction.UPLOAD_DOCUMENT)
        users = await get_all_users()
        await message.bot.send_chat_action(message.chat.id, ChatAction.UPLOAD_DOCUMENT)
        if not users:
            await message.answer("❌ Foydalanuvchilar topilmadi")
            return

        users_data = []
        for user in users:
            users_data.append(
                {
                    "ID": user.id,
                    "Telegram ID": user.user_id,
                    "Username": user.username or "—",
                    "To‘liq ismi": user.full_name or "—",
                    "Ro‘yxatdan o‘tgan vaqti": user.created_at.strftime(
                        "%d.%m.%Y %H:%M"
                    ),
                    "Oxirgi faolligi": user.last_active_at.strftime("%d.%m.%Y %H:%M"),
                    "Premium": "⭐ Ha" if user.is_premium else "❌ Yo‘q",
                }
            )

        df = pd.DataFrame(users_data)

        filename = f"users_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        filepath = f"media/files/{filename}"

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Foydalanuvchilar", index=False)
            workbook = writer.book
            worksheet = writer.sheets["Foydalanuvchilar"]

            for idx, col in enumerate(df.columns):
                max_length = (
                    max(df[col].astype(str).apply(len).max(), len(str(col))) + 2
                )
                worksheet.column_dimensions[
                    worksheet.cell(1, idx + 1).column_letter
                ].width = max_length

            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill(
                    start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
                )

        if os.path.exists(filepath):
            excel_file = FSInputFile(filepath)
            await message.bot.send_chat_action(
                message.chat.id, ChatAction.UPLOAD_DOCUMENT
            )
            await message.answer_document(
                document=excel_file,
                caption=(
                    f"📊 *Bot foydalanuvchilari ro‘yxati:*\n"
                    f"📅 *Sana:* {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
                    f"👥 *Jami:* {len(users):,} ta foydalanuvchi"
                ),
                parse_mode="Markdown",
            )

            try:
                os.remove(filepath)
            except FileNotFoundError:
                pass
        else:
            raise FileNotFoundError(f"Excel file not found at {filepath}")

    except Exception as e:
        print(f"Error creating Excel file: {e}")
        await message.answer("❌ Excel fayl yaratishda xatolik yuz berdi")


@router.message(AdminFilter(), F.text == CHANNEL_CONTROLLER)
async def channel_controller(message: Message):
    await message.answer(
        "🔧 *Kanal boshqaruv paneli*",
        parse_mode="Markdown",
        reply_markup=await channels_button(),
    )


@router.message(AdminFilter(), F.text == BACK)
async def go_back(message: Message):
    await message.answer(
        "🔙 *Admin panel*",
        parse_mode="Markdown",
        reply_markup=await menu_button(),
    )


@router.message(AdminFilter(), F.text == DATA_CONTROLLER)
async def channel_controller(message: Message):
    await message.answer(
        "🔧 *Ma'lumotlar boshqaruv paneli*",
        parse_mode="Markdown",
        reply_markup=await data_button(),
    )


@router.message(AdminFilter(), F.text == CATEGORY_CONTROLLER)
async def channel_controller(message: Message):
    await message.answer(
        "🔧 *Categoriyalar boshqaruv paneli*",
        parse_mode="Markdown",
        reply_markup=await category_button(),
    )
